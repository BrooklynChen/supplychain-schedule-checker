import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def format_report(df_file_name):
    wb = openpyxl.load_workbook(df_file_name)
    ws = wb.active

    orange_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
    yellow_fill = PatternFill(start_color='FEFDE8', end_color='FEFDE8', fill_type='solid')
    red_font = Font(color='C00000', bold=True)
    header_fill = PatternFill(start_color='16365C', end_color='16365C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    max_row = ws.max_row
    max_col = ws.max_column

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    for col in range(1, max_col + 1):
        header_cell = ws.cell(row=1, column=col)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[1].height = 30  # Set height to 30 points

    thin_border = Border(
        left=Side(style='thin', color='A6A6A6'),  # Gray color in RGB format
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='A6A6A6'),
        bottom=Side(style='thin', color='A6A6A6')
    )

    for row in range(1, max_row + 1):
        cell_a = ws[f'A{row}']
        cell_c = ws[f'C{row}']

        if cell_c.value == "Factory":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = orange_fill

        elif cell_a.value == "Not in Shipping Schedule":
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

        if cell_a.value == "Check manually":
            cell_a.font = red_font

        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Format columns I and J
    for row in range(1, max_row + 1):
        for col in ['I', 'J']:
            cell = ws[f'{col}{row}']
            if isinstance(cell.value, (int, float)):  # Check if the cell value is numeric
                cell.number_format = '#,##0'  # Format with comma and zero decimal places

    wb.save(df_file_name)